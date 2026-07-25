from __future__ import annotations

import json
import os
import re
import sqlite3
from datetime import UTC, datetime
from pathlib import Path
from typing import Any
from urllib.error import URLError
from urllib.request import urlopen

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
LOCAL_WORKBOOK_PATH = ROOT / "consumables" / "00000000_Consumables.xlsx"
DOWNLOADED_WORKBOOK_PATH = ROOT / "consumables" / "online_consumables.xlsx"
GOOGLE_SHEET_ID = "1r4f1ONTeUuzgldoYq1wlbfUGEKqKDLUx"
GOOGLE_WORKBOOK_URL = f"https://docs.google.com/spreadsheets/d/{GOOGLE_SHEET_ID}/export?format=xlsx"
DATA_DIR = ROOT / "data"
JSON_PATH = DATA_DIR / "consumables-seed.json"
SQLITE_PATH = DATA_DIR / "consumables.db"
LAB_HEAD_EMAIL = "umahajan@niper.ac.in"

HEADER_ALIASES = {
    "name": {"name", "consumables"},
    "manufacturer": {"company", "companyvendor", "manufacturer"},
    "catalogue_number": {"catno", "catalogno", "catlogno", "catalog", "catlog", "catlogno."},
    "quantity": {"quantity", "qty"},
    "approved_by": {"approvedby"},
    "mode_of_order": {"modeoforder"},
    "ordered_by": {"orderby", "orderedby"},
    "ordered_date": {"ordereddate", "orderdate", "purchasedate"},
    "project": {"project"},
    "date_of_issue": {"dateofissue"},
    "received_by": {"receivedby", "recievedby"},
    "stored_at": {"storedat", "location"},
    "status": {"status"},
}


def canonicalize_header(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    return re.sub(r"[^a-z0-9]+", "", text)


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        text = re.sub(r"\s+", " ", value).strip()
        return text or None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip() or None


def to_iso_date(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()

    raw = clean_text(value)
    if not raw:
        return None

    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d.%m.%y", "%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw, fmt).date().isoformat()
        except ValueError:
            continue

    return raw


def quantity_value(raw: str | None) -> float | None:
    if raw is None:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", raw)
    return float(match.group()) if match else None


def infer_status(record: dict[str, Any]) -> str:
    explicit = clean_text(record.get("status"))
    if explicit:
        return explicit
    if record.get("stored_at"):
        return "Available in storage"
    if record.get("ordered_date") or record.get("date_of_issue"):
        return "Ordered"
    return "Recorded"


def record_sort_key(record: dict[str, Any]) -> tuple[str, str, int]:
    latest = record.get("ordered_date") or record.get("date_of_issue") or ""
    issue = record.get("date_of_issue") or ""
    return latest, issue, int(record["source_row"])


def item_key(name: str | None, catalogue_number: str | None, manufacturer: str | None, category: str) -> str:
    parts = [name or "", catalogue_number or "", manufacturer or "", category]
    normalized = [re.sub(r"[^a-z0-9]+", "-", part.lower()).strip("-") for part in parts]
    return "::".join(part or "na" for part in normalized)


def resolve_workbook_source() -> tuple[Path, str]:
    source_preference = os.environ.get("CONSUMABLES_SOURCE", "online").strip().lower()
    if source_preference == "local":
        if not LOCAL_WORKBOOK_PATH.exists():
            raise FileNotFoundError(f"Local workbook not found: {LOCAL_WORKBOOK_PATH}")
        return LOCAL_WORKBOOK_PATH, str(LOCAL_WORKBOOK_PATH.relative_to(ROOT))

    try:
        with urlopen(GOOGLE_WORKBOOK_URL, timeout=30) as response:
            DOWNLOADED_WORKBOOK_PATH.parent.mkdir(parents=True, exist_ok=True)
            DOWNLOADED_WORKBOOK_PATH.write_bytes(response.read())
        return DOWNLOADED_WORKBOOK_PATH, GOOGLE_WORKBOOK_URL
    except URLError:
        if LOCAL_WORKBOOK_PATH.exists():
            return LOCAL_WORKBOOK_PATH, str(LOCAL_WORKBOOK_PATH.relative_to(ROOT))
        raise


def load_rows(workbook_path: Path) -> list[dict[str, Any]]:
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    rows: list[dict[str, Any]] = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        header_row = [canonicalize_header(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        column_map: dict[str, int] = {}
        for index, header in enumerate(header_row):
            for target, aliases in HEADER_ALIASES.items():
                if header in aliases and target not in column_map:
                    column_map[target] = index

        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(value not in (None, "") for value in row):
                continue

            record = {
                "category": sheet_name,
                "name": clean_text(row[column_map["name"]]) if "name" in column_map and column_map["name"] < len(row) else None,
                "manufacturer": clean_text(row[column_map["manufacturer"]]) if "manufacturer" in column_map and column_map["manufacturer"] < len(row) else None,
                "catalogue_number": clean_text(row[column_map["catalogue_number"]]) if "catalogue_number" in column_map and column_map["catalogue_number"] < len(row) else None,
                "quantity_text": clean_text(row[column_map["quantity"]]) if "quantity" in column_map and column_map["quantity"] < len(row) else None,
                "approved_by": clean_text(row[column_map["approved_by"]]) if "approved_by" in column_map and column_map["approved_by"] < len(row) else None,
                "mode_of_order": clean_text(row[column_map["mode_of_order"]]) if "mode_of_order" in column_map and column_map["mode_of_order"] < len(row) else None,
                "ordered_by": clean_text(row[column_map["ordered_by"]]) if "ordered_by" in column_map and column_map["ordered_by"] < len(row) else None,
                "ordered_date": to_iso_date(row[column_map["ordered_date"]]) if "ordered_date" in column_map and column_map["ordered_date"] < len(row) else None,
                "project": clean_text(row[column_map["project"]]) if "project" in column_map and column_map["project"] < len(row) else None,
                "date_of_issue": to_iso_date(row[column_map["date_of_issue"]]) if "date_of_issue" in column_map and column_map["date_of_issue"] < len(row) else None,
                "received_by": clean_text(row[column_map["received_by"]]) if "received_by" in column_map and column_map["received_by"] < len(row) else None,
                "stored_at": clean_text(row[column_map["stored_at"]]) if "stored_at" in column_map and column_map["stored_at"] < len(row) else None,
                "status": clean_text(row[column_map["status"]]) if "status" in column_map and column_map["status"] < len(row) else None,
                "source_sheet": sheet_name,
                "source_row": row_index,
            }

            if not record["name"]:
                continue

            record["status"] = infer_status(record)
            record["quantity_value"] = quantity_value(record["quantity_text"])
            record["item_key"] = item_key(record["name"], record["catalogue_number"], record["manufacturer"], sheet_name)
            record["search_text"] = " ".join(
                part for part in [record["name"], record["catalogue_number"], record["manufacturer"], sheet_name] if part
            ).lower()
            rows.append(record)

    return rows


def build_catalog(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    for row in rows:
        existing = grouped.get(row["item_key"])
        if existing is None or record_sort_key(row) >= record_sort_key(existing):
            grouped[row["item_key"]] = row
    return sorted(grouped.values(), key=lambda item: (item["name"].lower(), item["category"].lower()))


def write_json(catalog: list[dict[str, Any]], rows: list[dict[str, Any]], source_workbook: str) -> None:
    categories: dict[str, int] = {}
    for item in catalog:
        categories[item["category"]] = categories.get(item["category"], 0) + 1

    payload = {
        "generatedAt": datetime.now(UTC).replace(microsecond=0).isoformat().replace('+00:00', 'Z'),
        "labHeadEmail": LAB_HEAD_EMAIL,
        "sourceWorkbook": source_workbook,
        "summary": {
            "catalogItems": len(catalog),
            "orderRows": len(rows),
            "categories": categories,
        },
        "items": catalog,
        "orders": rows,
    }
    JSON_PATH.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def write_sqlite(catalog: list[dict[str, Any]], rows: list[dict[str, Any]]) -> None:
    if SQLITE_PATH.exists():
        SQLITE_PATH.unlink()

    connection = sqlite3.connect(SQLITE_PATH)
    cursor = connection.cursor()

    cursor.executescript(
        """
        CREATE TABLE catalog (
            item_key TEXT PRIMARY KEY,
            category TEXT NOT NULL,
            name TEXT NOT NULL,
            manufacturer TEXT,
            catalogue_number TEXT,
            quantity_text TEXT,
            quantity_value REAL,
            ordered_date TEXT,
            date_of_issue TEXT,
            stored_at TEXT,
            status TEXT,
            approved_by TEXT,
            ordered_by TEXT,
            received_by TEXT,
            project TEXT,
            search_text TEXT NOT NULL
        );

        CREATE TABLE order_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            item_key TEXT NOT NULL,
            category TEXT NOT NULL,
            name TEXT NOT NULL,
            manufacturer TEXT,
            catalogue_number TEXT,
            quantity_text TEXT,
            quantity_value REAL,
            ordered_date TEXT,
            date_of_issue TEXT,
            stored_at TEXT,
            status TEXT,
            approved_by TEXT,
            ordered_by TEXT,
            received_by TEXT,
            project TEXT,
            mode_of_order TEXT,
            source_sheet TEXT NOT NULL,
            source_row INTEGER NOT NULL,
            search_text TEXT NOT NULL,
            FOREIGN KEY (item_key) REFERENCES catalog(item_key)
        );

        CREATE TABLE purchase_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            requester_name TEXT NOT NULL,
            requester_email TEXT,
            requester_role TEXT,
            item_name TEXT NOT NULL,
            catalogue_number TEXT,
            manufacturer TEXT,
            requested_quantity TEXT,
            availability_status TEXT NOT NULL,
            request_notes TEXT,
            status TEXT NOT NULL DEFAULT 'pending_lab_head',
            created_at TEXT NOT NULL,
            confirmed_at TEXT,
            ordered_at TEXT,
            stored_at TEXT,
            lab_head_email TEXT NOT NULL
        );

        CREATE INDEX idx_catalog_search ON catalog(search_text);
        CREATE INDEX idx_order_history_item_key ON order_history(item_key);
        CREATE INDEX idx_order_history_search ON order_history(search_text);
        """
    )

    cursor.executemany(
        """
        INSERT INTO catalog (
            item_key, category, name, manufacturer, catalogue_number, quantity_text, quantity_value,
            ordered_date, date_of_issue, stored_at, status, approved_by, ordered_by, received_by,
            project, search_text
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        [
            (
                item["item_key"], item["category"], item["name"], item["manufacturer"], item["catalogue_number"],
                item["quantity_text"], item["quantity_value"], item["ordered_date"], item["date_of_issue"],
                item["stored_at"], item["status"], item["approved_by"], item["ordered_by"], item["received_by"],
                item["project"], item["search_text"],
            )
            for item in catalog
        ],
    )

    cursor.executemany(
        """
        INSERT INTO order_history (
            item_key, category, name, manufacturer, catalogue_number, quantity_text, quantity_value,
            ordered_date, date_of_issue, stored_at, status, approved_by, ordered_by, received_by,
            project, mode_of_order, source_sheet, source_row, search_text
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        [
            (
                row["item_key"], row["category"], row["name"], row["manufacturer"], row["catalogue_number"],
                row["quantity_text"], row["quantity_value"], row["ordered_date"], row["date_of_issue"],
                row["stored_at"], row["status"], row["approved_by"], row["ordered_by"], row["received_by"],
                row["project"], row["mode_of_order"], row["source_sheet"], row["source_row"], row["search_text"],
            )
            for row in rows
        ],
    )

    connection.commit()
    connection.close()


def main() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    workbook_path, source_workbook = resolve_workbook_source()
    rows = load_rows(workbook_path)
    catalog = build_catalog(rows)
    write_json(catalog, rows, source_workbook)
    write_sqlite(catalog, rows)
    print(f"Wrote {len(catalog)} catalog items and {len(rows)} order history rows.")
    print(f"Source: {source_workbook}")
    print(f"JSON: {JSON_PATH.relative_to(ROOT)}")
    print(f"SQLite: {SQLITE_PATH.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
